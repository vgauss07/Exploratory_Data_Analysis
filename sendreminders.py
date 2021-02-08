import openpyxl, smtplib, sys

# Open the spreadsheet and get the latest dues status.

wb = openpyxl.load_workbook("Sala 'Amore Payment Report.xlsx")
sheet = wb['Payment Status']

lastCol = sheet.max_column
latestMonth = sheet.cell(row=1, column=lastCol).value

# TODO: Check each member's payment status.
# TODO: Log in to email account
# TODO: Send out reminder emails.


unpaidMembers = {}
for r in range(2, sheet.max_row + 1):
    payment = sheet.cell(row=r, column=lastCol).value
    if payment != 'Paid':
        name = sheet.cell(row=r, column=1).value
        email = sheet.cell(row=r, column=2).value
        unpaidMembers[name] = email

# send customized emails
smtpObj = smtplib.SMTP('smtp.gmail.com', 587)
smtpObj.ehlo()
smtpObj.starttls()
smtpObj.login(' my_email_address@gmail.com ', sys.argv[1])

# send out reminder emails.
for name, email in unpaidMembers.items():
    body = "Subject: %s dues unpaid. \nDear %s, \nRecords show that you have not" \
           "paid dues for %s. Please make this payment as soon as possible. Thank you!" % (
           latestMonth, name, latestMonth)
    print('Sending email to %s...' % email)
    sendmailStatus = smtpObj.sendmail('my_email_address@gmail.com', email, body)

    if sendmailStatus != {}:
        print('There was a problem sending email to %s:%s' % (email, sendmailStatus))
    smtpObj.quit()
